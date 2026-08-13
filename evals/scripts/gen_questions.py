"""题集生成器（防污染机制 4：答案值可执行复算）。

核心约定：**F1 / F3 / F4 三类题的答案值不是手填的，而是 SQL 从已发布语料产物中执行得到。**
题面与答案同时生成，且每题记录 source_sql，使答案值可被独立复算。这能减少手工抄写答案
造成的错误，但不会消除标注错误：题面到 SQL 的语义映射、模板和数据构造仍由作者设计，
必须由审计、扰动对照和外部基准共同约束。

数据来源刻意选择**已发布的产物**（SQLite / XLSX）而非内存中的中间结果，
以保证题集与仓库里实际提交的语料严格一致。

F2 跨源冲突 / F5 多轮钻取 / F6 拒答题为手工设计，答案同样由发布资产或明确的不可答契约
复核，见 questions/manual/。

用法::

    python evals/scripts/gen_questions.py
"""

from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))

from perturb import perturbed_keys

ROOT = Path(__file__).resolve().parents[1]
NOVELS = ROOT / "novels"
QUESTIONS = ROOT / "questions"

# 每部各题型的目标题数（合计 60/部，两部 120）
QUOTA = {"F1": 16, "F3": 14, "F4": 12}  # F2=8 / F5=6 / F6=4 为手工题


# ---------------------------------------------------------------- 载入已发布产物

def load_xiyouji() -> sqlite3.Connection:
    """直接使用已发布的 SQLite 产物。"""
    src = sqlite3.connect(NOVELS / "xiyouji" / "sql" / "xiyouji.db")
    mem = sqlite3.connect(":memory:")
    src.backup(mem)
    src.close()
    mem.row_factory = sqlite3.Row
    return mem


def load_shuihu() -> sqlite3.Connection:
    """从已发布的 XLSX 产物读回 → 内存 SQLite。

    读取时必须正确处理三处脏结构（合并单元格标题行、表头空格、年份存为文本），
    这本身就是对语料"确实很脏"的一次验证。
    """
    from openpyxl import load_workbook

    wb = load_workbook(NOVELS / "shuihu" / "xlsx" / "liangshan.xlsx", data_only=True)
    ws = wb["头领名录"]
    rows_iter = ws.iter_rows(min_row=2, values_only=True)
    header = [str(h).strip() for h in next(rows_iter)]  # 表头在第 2 行，且带空格
    assert header[0] == "座次", f"表头解析异常: {header}"

    con = sqlite3.connect(":memory:")
    con.row_factory = sqlite3.Row
    con.execute("""
        CREATE TABLE liangshan (
          rank INTEGER PRIMARY KEY, star_name TEXT, nickname TEXT, name TEXT,
          camp TEXT, role_group TEXT, home_town TEXT,
          join_year INTEGER, merit_count INTEGER, outcome TEXT
        )""")
    def as_int(v: Any) -> int:
        return int(str(v).strip())

    for r in rows_iter:
        if r[0] is None:
            continue
        con.execute("INSERT INTO liangshan VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (as_int(r[0]), r[1], r[2], r[3], r[4], r[5], r[6],
                     as_int(r[7]), as_int(r[8]), r[9]))  # 年份存为文本 → 整数
    con.commit()
    return con


# ---------------------------------------------------------------- 题目构造

def q(con: sqlite3.Connection, qid: str, book: str, qtype: str, question: str,
      sql: str, params: tuple = (), *, kind: str = "scalar",
      field_class: str = "curated", targets_perturbed_field: bool = False,
      template_id: str, tolerance: float = 0.0) -> dict[str, Any]:
    """执行 SQL 得到 ground truth，构造一道题。

    三个统计字段随资产一起冻结，不由下游 loader 猜：

    * ``template_id``：同一 NL 模板的不同实例共享它。它是 contamination diagnostic 的
      cluster key，必须来自生成器而不是事后从题面正则推断。
    * ``independence_unit_id`` / ``unit_kind``：确认性统计的分析单位。单轮题一题一个
      unit；F5 由手工题生成器写成整条 chain 一个 unit。
    * ``targets_perturbed_field``：这道题问的字段是否被扰动过。它是**题目属性**，
      不是数据集分区——发布语料本身就是 perturbed snapshot，见 docs/BENCHMARK.md §5。
    """
    cur = con.execute(sql, params)
    fetched = cur.fetchall()
    if kind == "scalar":
        gt: Any = fetched[0][0] if fetched else None
    elif kind == "number":
        gt = round(float(fetched[0][0]), 4) if fetched and fetched[0][0] is not None else None
    elif kind == "set":
        gt = sorted(str(r[0]) for r in fetched)
    else:  # pragma: no cover
        raise ValueError(kind)

    return {
        "id": qid, "book": book, "type": qtype,
        "template_id": template_id,
        "independence_unit_id": qid,
        "unit_kind": "single",
        "question": question,
        "ground_truth": gt,
        "answer_kind": kind,
        "scoring": {"mode": {"scalar": "exact", "number": "numeric", "set": "set"}[kind],
                    "tolerance": tolerance},
        "trace_required": True,
        "source_sql": " ".join(sql.split()),
        "source_params": list(params),
        "field_class": field_class,
        "targets_perturbed_field": targets_perturbed_field,
    }


PREFIX = "按本系统记载，"   # 统一前缀：把问题锚定到数据而非世界知识


def build_xiyouji(con: sqlite3.Connection) -> list[dict]:
    pk = perturbed_keys("xiyouji")
    out: list[dict] = []
    n = 0

    def nid(t: str) -> str:
        nonlocal n
        n += 1
        return f"XY-{t}-{n:03d}"

    # ---- F1 单源查询（16）----
    # 元组第一位是 template_id：同一模板的多个实例共享它，供 cluster bootstrap 使用。
    f1_specs: list[tuple[str, str, str, tuple, str, str]] = []
    for no in (3, 17, 29, 44, 58, 70):
        f1_specs.append(("xy.f1.nan_name_by_no", f"{PREFIX}第 {no} 难的名称是什么？",
                         "SELECT nan_name FROM nan WHERE nan_no=?", (no,), "scalar", "canonical"))
    for no in (11, 33, 55, 76):
        f1_specs.append(("xy.f1.opponent_by_no", f"{PREFIX}第 {no} 难的对手是谁？",
                         "SELECT opponent FROM nan WHERE nan_no=?", (no,), "scalar", "curated"))
    for no in (18, 60):
        f1_specs.append(("xy.f1.helper_by_no", f"{PREFIX}第 {no} 难的援手是谁？",
                         "SELECT helper FROM nan WHERE nan_no=?", (no,), "scalar", "curated"))
    for no in (22, 65):
        f1_specs.append(("xy.f1.location_by_no", f"{PREFIX}第 {no} 难发生在哪里？",
                         "SELECT location FROM nan WHERE nan_no=?", (no,), "scalar", "curated"))
    for no in (7, 50):
        f1_specs.append(("xy.f1.duration_by_no", f"{PREFIX}第 {no} 难历时多少天？",
                         "SELECT duration_days FROM nan WHERE nan_no=?", (no,), "scalar", "synthetic"))
    for tpl, question, sql, params, kind, fc in f1_specs[:QUOTA["F1"]]:
        out.append(q(con, nid("F1"), "xiyouji", "F1", question, sql, params,
                     kind=kind, field_class=fc, template_id=tpl))

    # ---- F3 聚合统计（14）----
    # 每道 F3 是一次性聚合口径，各自成一个 template。
    f3: list[tuple[str, str, str, tuple, str, str]] = [
        ("xy.f3.water_count", f"{PREFIX}八十一难中「水难」类共有多少难？",
         "SELECT COUNT(*) FROM nan WHERE category='水难'", (), "scalar", "curated"),
        ("xy.f3.demon_share", f"{PREFIX}「妖魔」类占全部难数的比例是多少？（保留 4 位小数）",
         "SELECT CAST(SUM(category='妖魔') AS REAL)/COUNT(*) FROM nan", (), "number", "curated"),
        ("xy.f3.avg_difficulty", f"{PREFIX}所有难的平均难度评分是多少？（保留 4 位小数）",
         "SELECT AVG(difficulty_score) FROM nan", (), "number", "synthetic"),
        ("xy.f3.longest_duration_no", f"{PREFIX}历时天数最长的是第几难？",
         "SELECT nan_no FROM nan ORDER BY duration_days DESC, nan_no LIMIT 1", (), "scalar", "synthetic"),
        ("xy.f3.guanyin_helper_count", f"{PREFIX}援手为「观音」的难共有多少条？",
         "SELECT COUNT(*) FROM nan WHERE helper='观音'", (), "scalar", "curated"),
        ("xy.f3.top_location", f"{PREFIX}出现次数最多的发生地是哪里？",
         "SELECT location FROM nan GROUP BY location ORDER BY COUNT(*) DESC, location LIMIT 1",
         (), "scalar", "curated"),
        ("xy.f3.chapter_range_count", f"{PREFIX}近似章回在 40 至 60 之间（含）的难共有多少条？",
         "SELECT COUNT(*) FROM nan WHERE chapter_hint BETWEEN 40 AND 60", (), "scalar", "curated"),
        ("xy.f3.total_duration", f"{PREFIX}全部难的历时天数总和是多少？",
         "SELECT SUM(duration_days) FROM nan", (), "scalar", "synthetic"),
        ("xy.f3.difficulty_10_names", f"{PREFIX}难度评分为 10 的难有哪些？请列出难名。",
         "SELECT nan_name FROM nan WHERE difficulty_score=10", (), "set", "synthetic"),
        ("xy.f3.renhuo_top_difficulty", f"{PREFIX}「人祸」类中难度评分最高的是哪一难？",
         "SELECT nan_name FROM nan WHERE category='人祸' ORDER BY difficulty_score DESC, nan_no LIMIT 1",
         (), "scalar", "curated"),
        ("xy.f3.distinct_location_count", f"{PREFIX}共有多少个不同的发生地？",
         "SELECT COUNT(DISTINCT location) FROM nan", (), "scalar", "curated"),
        ("xy.f3.null_opponent_count", f"{PREFIX}对手字段为空（未记载对手）的难有多少条？",
         "SELECT COUNT(*) FROM nan WHERE opponent IS NULL OR opponent=''", (), "scalar", "curated"),
        ("xy.f3.shenfo_avg_duration", f"{PREFIX}「神佛试炼」类的平均历时天数是多少？（保留 4 位小数）",
         "SELECT AVG(duration_days) FROM nan WHERE category='神佛试炼'", (), "number", "synthetic"),
        ("xy.f3.demon_character_count", f"{PREFIX}人物表中阵营为「妖魔」的角色共有多少个？",
         "SELECT COUNT(*) FROM characters WHERE camp='妖魔'", (), "scalar", "curated"),
    ]
    for tpl, question, sql, params, kind, fc in f3[:QUOTA["F3"]]:
        out.append(q(con, nid("F3"), "xiyouji", "F3", question, sql, params,
                     kind=kind, field_class=fc, template_id=tpl,
                     tolerance=1e-4 if kind == "number" else 0.0))

    # ---- F4 扰动反事实（12）：只问被扰动过的记录/字段 ----
    f4: list[tuple[str, str, str, tuple, str]] = []
    for no in sorted(pk.get("opponent", ()))[:6]:
        f4.append(("xy.f4.opponent_by_no",
                   f"{PREFIX}第 {no} 难的对手是谁？（请以本系统数据为准，不要依据你已有的印象）",
                   "SELECT opponent FROM nan WHERE nan_no=?", (no,), "scalar"))
    for no in sorted(pk.get("nan_name", ()))[:3]:
        f4.append(("xy.f4.nan_name_by_no",
                   f"{PREFIX}第 {no} 难的名称是什么？（请以本系统数据为准）",
                   "SELECT nan_name FROM nan WHERE nan_no=?", (no,), "scalar"))
    for no in sorted(pk.get("helper", ()))[:1]:
        f4.append(("xy.f4.helper_by_no", f"{PREFIX}第 {no} 难由谁相助解厄？",
                   "SELECT helper FROM nan WHERE nan_no=?", (no,), "scalar"))
    for no in sorted(pk.get("location", ()))[:1]:
        f4.append(("xy.f4.location_by_no", f"{PREFIX}第 {no} 难发生在何处？",
                   "SELECT location FROM nan WHERE nan_no=?", (no,), "scalar"))
    f4.append(("xy.f4.tianzai_count",
               f"{PREFIX}「天灾」类共有多少难？（注意：类别归属以本系统数据为准）",
               "SELECT COUNT(*) FROM nan WHERE category='天灾'", (), "scalar"))
    for tpl, question, sql, params, kind in f4[:QUOTA["F4"]]:
        out.append(q(con, nid("F4"), "xiyouji", "F4", question, sql, params,
                     kind=kind, field_class="canonical", template_id=tpl,
                     targets_perturbed_field=True))
    return out


def build_shuihu(con: sqlite3.Connection) -> list[dict]:
    pk = perturbed_keys("shuihu")
    out: list[dict] = []
    n = 0

    def nid(t: str) -> str:
        nonlocal n
        n += 1
        return f"SH-{t}-{n:03d}"

    # ---- F1（16）----
    f1: list[tuple[str, str, str, tuple, str, str]] = []
    for r in (7, 20, 41, 66, 88, 102):
        f1.append(("sh.f1.name_by_rank", f"{PREFIX}座次第 {r} 位的姓名是什么？",
                   "SELECT name FROM liangshan WHERE rank=?", (r,), "scalar", "canonical"))
    for r in (9, 34, 59, 95):
        f1.append(("sh.f1.nickname_by_rank", f"{PREFIX}座次第 {r} 位的绰号是什么？",
                   "SELECT nickname FROM liangshan WHERE rank=?", (r,), "scalar", "canonical"))
    for r in (12, 77):
        f1.append(("sh.f1.star_name_by_rank", f"{PREFIX}座次第 {r} 位的星号是什么？",
                   "SELECT star_name FROM liangshan WHERE rank=?", (r,), "scalar", "canonical"))
    for r in (5, 48):
        f1.append(("sh.f1.home_town_by_rank", f"{PREFIX}座次第 {r} 位的籍贯是哪里？",
                   "SELECT home_town FROM liangshan WHERE rank=?", (r,), "scalar", "synthetic"))
    for r in (30, 81):
        f1.append(("sh.f1.merit_by_rank", f"{PREFIX}座次第 {r} 位的战功数是多少？",
                   "SELECT merit_count FROM liangshan WHERE rank=?", (r,), "scalar", "synthetic"))
    for tpl, question, sql, params, kind, fc in f1[:QUOTA["F1"]]:
        out.append(q(con, nid("F1"), "shuihu", "F1", question, sql, params,
                     kind=kind, field_class=fc, template_id=tpl))

    # ---- F3（14）----
    f3: list[tuple[str, str, str, tuple, str, str]] = [
        ("sh.f3.tiangang_count", f"{PREFIX}天罡星共有多少人？",
         "SELECT COUNT(*) FROM liangshan WHERE camp='天罡星'", (), "scalar", "canonical"),
        ("sh.f3.tiger_group_count", f"{PREFIX}「马军五虎将」职司分组共有多少人？",
         "SELECT COUNT(*) FROM liangshan WHERE role_group='马军五虎将'", (), "scalar", "curated"),
        ("sh.f3.navy_group_count", f"{PREFIX}「水军头领」职司分组共有多少人？",
         "SELECT COUNT(*) FROM liangshan WHERE role_group='水军头领'", (), "scalar", "curated"),
        ("sh.f3.total_merit", f"{PREFIX}全部头领的战功数总和是多少？",
         "SELECT SUM(merit_count) FROM liangshan", (), "scalar", "synthetic"),
        ("sh.f3.top_merit_name", f"{PREFIX}战功数最高的头领姓名是什么？",
         "SELECT name FROM liangshan ORDER BY merit_count DESC, rank LIMIT 1", (), "scalar", "synthetic"),
        ("sh.f3.yuncheng_count", f"{PREFIX}籍贯为「郓城县」的头领共有多少人？",
         "SELECT COUNT(*) FROM liangshan WHERE home_town='郓城县'", (), "scalar", "synthetic"),
        ("sh.f3.kia_share", f"{PREFIX}结局为「征战阵亡」的头领占全部头领的比例是多少？（保留 4 位小数）",
         "SELECT CAST(SUM(outcome='征战阵亡') AS REAL)/COUNT(*) FROM liangshan", (), "number", "synthetic"),
        ("sh.f3.min_join_year", f"{PREFIX}上山年份最早是哪一年？",
         "SELECT MIN(join_year) FROM liangshan", (), "scalar", "synthetic"),
        ("sh.f3.disha_avg_merit", f"{PREFIX}地煞星的平均战功数是多少？（保留 4 位小数）",
         "SELECT AVG(merit_count) FROM liangshan WHERE camp='地煞星'", (), "number", "synthetic"),
        ("sh.f3.distinct_role_group_count", f"{PREFIX}共有多少个不同的职司分组？",
         "SELECT COUNT(DISTINCT role_group) FROM liangshan", (), "scalar", "curated"),
        ("sh.f3.largest_role_group", f"{PREFIX}人数最多的职司分组是哪一个？",
         "SELECT role_group FROM liangshan GROUP BY role_group ORDER BY COUNT(*) DESC, role_group LIMIT 1",
         (), "scalar", "curated"),
        ("sh.f3.tiangang_merit_gt50", f"{PREFIX}天罡星中战功数超过 50 的有多少人？",
         "SELECT COUNT(*) FROM liangshan WHERE camp='天罡星' AND merit_count>50", (), "scalar", "synthetic"),
        ("sh.f3.top_outcome", f"{PREFIX}结局的分布中，出现次数最多的结局是什么？",
         "SELECT outcome FROM liangshan GROUP BY outcome ORDER BY COUNT(*) DESC, outcome LIMIT 1",
         (), "scalar", "synthetic"),
        ("sh.f3.join_year_1110_names", f"{PREFIX}上山年份为 1110 年的头领有哪些？请列出姓名。",
         "SELECT name FROM liangshan WHERE join_year=1110", (), "set", "synthetic"),
    ]
    for tpl, question, sql, params, kind, fc in f3[:QUOTA["F3"]]:
        out.append(q(con, nid("F3"), "shuihu", "F3", question, sql, params,
                     kind=kind, field_class=fc, template_id=tpl,
                     tolerance=1e-4 if kind == "number" else 0.0))

    # ---- F4 扰动反事实（12）----
    f4: list[tuple[str, str, str, tuple, str]] = []
    for r in sorted(pk.get("name", ())):
        f4.append(("sh.f4.name_by_rank",
                   f"{PREFIX}座次第 {r} 位是谁？（请以本系统数据为准，不要依据你已有的印象）",
                   "SELECT name FROM liangshan WHERE rank=?", (r,), "scalar"))
        f4.append(("sh.f4.nickname_by_rank", f"{PREFIX}座次第 {r} 位的绰号是什么？",
                   "SELECT nickname FROM liangshan WHERE rank=?", (r,), "scalar"))
    for r in sorted(pk.get("nickname", set()) - pk.get("name", set())):
        f4.append(("sh.f4.nickname_by_rank", f"{PREFIX}座次第 {r} 位的绰号是什么？（以本系统数据为准）",
                   "SELECT nickname FROM liangshan WHERE rank=?", (r,), "scalar"))
    for r in sorted(pk.get("role_group", ())):
        f4.append(("sh.f4.role_group_by_rank", f"{PREFIX}座次第 {r} 位属于哪个职司分组？",
                   "SELECT role_group FROM liangshan WHERE rank=?", (r,), "scalar"))
    f4 += [
        ("sh.f4.rank_by_nickname", f"{PREFIX}绰号为「铁和尚」的头领座次是第几位？",
         "SELECT rank FROM liangshan WHERE nickname='铁和尚'", (), "scalar"),
        ("sh.f4.tiger_group_names", f"{PREFIX}「马军五虎将」都有谁？请列出姓名。",
         "SELECT name FROM liangshan WHERE role_group='马军五虎将'", (), "set"),
        ("sh.f4.rank_by_name", f"{PREFIX}姓名为「吴用」的头领座次是第几位？",
         "SELECT rank FROM liangshan WHERE name='吴用'", (), "scalar"),
        ("sh.f4.rank_by_name", f"{PREFIX}姓名为「徐宁」的头领座次是第几位？",
         "SELECT rank FROM liangshan WHERE name='徐宁'", (), "scalar"),
    ]
    for tpl, question, sql, params, kind in f4[:QUOTA["F4"]]:
        out.append(q(con, nid("F4"), "shuihu", "F4", question, sql, params,
                     kind=kind, field_class="canonical", template_id=tpl,
                     targets_perturbed_field=True))
    return out


# ---------------------------------------------------------------- 出口

def main() -> None:
    QUESTIONS.mkdir(parents=True, exist_ok=True)
    total = 0
    for book, loader, builder in (("xiyouji", load_xiyouji, build_xiyouji),
                                  ("shuihu", load_shuihu, build_shuihu)):
        con = loader()
        items = builder(con)
        con.close()

        # ground truth 自校验：任何 None 都说明模板写错了表/列
        bad = [i["id"] for i in items if i["ground_truth"] in (None, [], "")]
        if bad:
            raise SystemExit(f"[gen] {book} 以下题目的 ground truth 为空，模板有误: {bad}")

        # 统计字段自校验：缺 template_id 会让 contamination diagnostic 的 cluster key 落空，
        # 缺 independence_unit_id 会让配对检验退化成伪重复，两者都必须在生成期挡住。
        for key in ("template_id", "independence_unit_id", "unit_kind"):
            missing = [i["id"] for i in items if not i.get(key)]
            if missing:
                raise SystemExit(f"[gen] {book} 以下题目缺少 {key}: {missing}")

        path = QUESTIONS / f"{book}.jsonl"
        with path.open("w", encoding="utf-8") as f:
            for it in items:
                f.write(json.dumps(it, ensure_ascii=False) + "\n")

        counts: dict[str, int] = {}
        for it in items:
            counts[it["type"]] = counts.get(it["type"], 0) + 1
        total += len(items)
        n_tpl = len({it["template_id"] for it in items})
        print(f"[gen] {book}: {len(items)} 题（模板）  {counts}  template={n_tpl}")

    print(f"[gen] 模板题合计 {total} 题 → evals/questions/")
    print("[gen] 手工题（F2/F5/F6 共 36 题）见 evals/questions/manual/")


if __name__ == "__main__":
    main()
