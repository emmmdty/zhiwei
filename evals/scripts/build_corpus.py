"""语料构建器：由基础数据 + 扰动声明重建全部四种数据形态。

数据即代码：改一行 ``data/*.py``，跑一次本脚本，SQLite / CSV / XLSX / PDF+MD 四种形态同步重建。
产物全部入库（见 .gitignore 的 evals 例外规则），保证任何人 clone 后无需跑脚本也能复现跑分。

用法::

    python evals/scripts/build_corpus.py
"""

from __future__ import annotations

import csv
import hashlib
import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from data import shuihu as shuihu_data
from data import xiyouji as xiyouji_data
from perturb import CONFLICTS, apply_perturbations, conflict

ROOT = Path(__file__).resolve().parents[1]
NOVELS = ROOT / "novels"


# ---------------------------------------------------------------- 工具

def _write(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


# 固定时间戳：XLSX 是内嵌 mtime 的 ZIP，PDF 内嵌 /CreationDate 与文档 ID，
# 两者默认都会让"重跑得到字节一致的语料"这个承诺落空。
FIXED_TS = (1980, 1, 1, 0, 0, 0)
FIXED_DT = datetime(2026, 1, 1, 0, 0, 0)


def _normalize_zip(path: Path) -> None:
    """重写 ZIP 归档，使 XLSX 字节可复现。

    需要归一三处，缺一不可：
    1. ZIP 条目的 mtime（默认取当前时间）
    2. 条目顺序（按文件名排序）
    3. ``docProps/core.xml`` 里的 ``dcterms:modified`` —— **openpyxl 在 save() 时会用当前
       UTC 秒覆盖调用方设置的值**。它只有秒级精度，所以只有跨秒的两次构建才会产生差异，
       表现为"偶发的校验和对不上"，是最难排查的一类不可复现问题。
    """
    import re
    import zipfile

    fixed = FIXED_DT.strftime("%Y-%m-%dT%H:%M:%SZ")

    def scrub(name: str, data: bytes) -> bytes:
        if name != "docProps/core.xml":
            return data
        return re.sub(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                      rb"\g<1>" + fixed.encode() + rb"\g<2>", data)

    with zipfile.ZipFile(path) as src:
        entries = [(i.filename, i.external_attr, scrub(i.filename, src.read(i.filename)))
                   for i in sorted(src.infolist(), key=lambda x: x.filename)]
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for name, attr, data in entries:
            info = zipfile.ZipInfo(name, date_time=FIXED_TS)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = attr
            out.writestr(info, data)
    tmp.replace(path)


# ---------------------------------------------------------------- 西游记

def build_xiyouji() -> dict:
    base = xiyouji_data.rows()
    rows, manifest = apply_perturbations("xiyouji", base)
    chars = xiyouji_data.character_rows()
    out = NOVELS / "xiyouji"

    # --- SQLite ---
    db_path = out / "sql" / "xiyouji.db"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db_path.unlink(missing_ok=True)
    con = sqlite3.connect(db_path)
    con.executescript("""
        CREATE TABLE nan (
          nan_no           INTEGER PRIMARY KEY,   -- 难序号
          nan_name         TEXT NOT NULL,         -- 难名
          category         TEXT NOT NULL,         -- 类别（本项目自定义口径）
          location         TEXT,                  -- 发生地
          opponent         TEXT,                  -- 对手
          helper           TEXT,                  -- 援手
          chapter_hint     INTEGER,               -- 近似章回（非考据依据）
          duration_days    INTEGER,               -- 历时天数（虚构列）
          difficulty_score INTEGER                -- 难度评分 1-10（虚构列）
        );
        CREATE TABLE characters (
          name      TEXT PRIMARY KEY,  -- 姓名
          role      TEXT,              -- 身份
          weapon    TEXT,              -- 兵器/法宝
          first_nan INTEGER,           -- 首次登场难序号
          camp      TEXT               -- 阵营
        );
    """)
    con.executemany(
        "INSERT INTO nan VALUES (:nan_no,:nan_name,:category,:location,:opponent,"
        ":helper,:chapter_hint,:duration_days,:difficulty_score)", rows)
    con.executemany(
        "INSERT INTO characters VALUES (:name,:role,:weapon,:first_nan,:camp)", chars)
    con.commit()
    con.close()

    # --- CSV（同一数据的第二形态）---
    def dump_csv(path: Path, data: list[dict]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(data[0].keys()))
            w.writeheader()
            w.writerows(data)

    dump_csv(out / "csv" / "nan.csv", rows)
    dump_csv(out / "csv" / "characters.csv", chars)

    # --- Markdown 改编节选（第三形态，含受控跨源冲突）---
    by_no = {r["nan_no"]: r for r in rows}
    n_water = sum(1 for r in rows if r["category"] == "水难")
    # 四处受控冲突：文档值与表值刻意不一致，正确行为是报告分歧
    conflict_36 = conflict("xiyouji", "XY-C01")["doc_value"]          # 地点
    conflict_wukong = conflict("xiyouji", "XY-C02")["doc_value"]      # 首登难序号
    conflict_water = n_water + 1                                      # XY-C03 水难数
    conflict_81 = conflict("xiyouji", "XY-C04")["doc_value"]          # 第81难对手

    lines = [
        "# 取经路上八十一难 · 改编节选",
        "",
        "> 本文为**本项目新创作的改编文本**，非原著引文，无版权风险。",
        "> 文中记载与结构化表格**存在两处刻意保留的分歧**（见语料 README 的受控冲突章节），",
        "> 用于考验系统在跨源矛盾下是否如实报告分歧，而不是擅自择一。",
        "",
        "## 卷一 · 起因",
        "",
        (f"金蝉子遭贬，是为第一难。此后出胎、抛江、报冤三事相继，皆属宿因，"
         f"共历时 {sum(by_no[i]['duration_days'] for i in (1, 2, 3, 4))} 日。"),
        "",
        f"孙悟空于第 {conflict_wukong} 难现身，随行护法。",
        "",
        "## 卷二 · 途中诸难",
        "",
    ]
    for cat in ["妖魔", "水难", "人祸", "天灾", "神佛试炼"]:
        sel = [r for r in rows if r["category"] == cat]
        names = "、".join(r["nan_name"] for r in sel[:6])
        # XY-C03：水难一类的数目刻意写成表中值 + 1
        declared = conflict_water if cat == "水难" else len(sel)
        lines.append(f"### {cat}（共 {declared} 难）")
        lines.append("")
        lines.append(f"计有：{names} 等。其中难度评分最高者为"
                     f"「{max(sel, key=lambda r: r['difficulty_score'])['nan_name']}」。")
        lines.append("")

    lines += [
        "## 卷三 · 通天河",
        "",
        f"第 36 难「{by_no[36]['nan_name']}」，起于 {conflict_36} 之畔。",
        f"其时妖王名唤{by_no[36]['opponent']}，据水为患，师徒受阻。",
        f"至第 38 难方由{by_no[38]['helper']}以竹篮收之。",
        "",
        "## 卷四 · 终局",
        "",
        (f"第 81 难「{by_no[81]['nan_name']}」，{conflict_81}负经渡河，"
         f"经卷落水，晾晒于石上，遂成圆满。"),
        "",
    ]
    _write(out / "docs" / "xiyouji_notes.md", "\n".join(lines))

    _write(out / "perturbation_manifest.json",
           json.dumps(manifest, ensure_ascii=False, indent=2) + "\n")

    return {"book": "xiyouji", "rows": len(rows), "characters": len(chars),
            "db_sha256": _sha256(db_path), "manifest": manifest}


# ---------------------------------------------------------------- 水浒传

def build_shuihu() -> dict:
    base = shuihu_data.rows()
    rows, manifest = apply_perturbations("shuihu", base)
    out = NOVELS / "shuihu"

    # --- Excel（第四形态，故意制造三种真实世界的脏结构）---
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active or wb.create_sheet()
    ws.title = "头领名录"

    # 陷阱 1：标题占据首行并合并单元格 —— 真实表头在第 2 行，天真的解析器会读错
    ws["A1"] = "梁山泊一百单八将名录（本项目整理版）"
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 陷阱 2：中文表头带前后空格与全角括号
    headers = [" 座次 ", "星号", "绰号", " 姓名", "阵营（天罡/地煞）",
               "职司分组", "籍贯", "上山年份", " 战功数 ", "结局"]
    ws.append(headers)
    for c in ws[2]:
        c.font = Font(bold=True)

    for r in rows:
        ws.append([
            r["rank"], r["star_name"], r["nickname"], r["name"], r["camp"],
            r["role_group"], r["home_town"],
            # 陷阱 3：上山年份存为文本而非数字
            str(r["join_year"]),
            r["merit_count"], r["outcome"],
        ])

    # 第二个 sheet：职司汇总（考验多 sheet 定位）
    ws2 = wb.create_sheet("职司汇总")
    ws2.append(["职司分组", "人数", "战功数合计"])
    groups: dict[str, list[dict]] = {}
    for r in rows:
        groups.setdefault(r["role_group"], []).append(r)
    for g, members in sorted(groups.items()):
        ws2.append([g, len(members), sum(m["merit_count"] for m in members)])

    xlsx_path = out / "xlsx" / "liangshan.xlsx"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.properties.created = FIXED_DT      # docProps/core.xml 里的创建/修改时间
    wb.properties.modified = FIXED_DT
    wb.save(xlsx_path)
    _normalize_zip(xlsx_path)

    # --- 招安文档（Markdown + PDF）---
    by_rank = {r["rank"]: r for r in rows}
    wusong = next(r for r in rows if r["name"] == "武松")
    n_navy = sum(1 for r in rows if r["role_group"] == "水军头领")
    # 四处受控冲突（SH-C01..C04）：文档值与表值刻意不一致
    conflict_merit = wusong["merit_count"] + 7
    conflict_tiger = conflict("shuihu", "SH-C02")["doc_value"]
    conflict_navy = n_navy - 2
    conflict_year = min(r["join_year"] for r in rows) - 3

    md = [
        "# 梁山招安始末 · 改编节选",
        "",
        "> 本文为**本项目新创作的改编文本**，非原著引文。",
        "> 文中「武松战功数」与 Excel 名录**刻意不一致**（受控冲突 SH-C01），",
        "> 正确行为是报告分歧并给出两处出处。",
        "",
        "## 一 · 排座次",
        "",
        (f"天罡三十六星，地煞七十二星，共一百单八人。座次首者为"
         f"{by_rank[1]['nickname']}{by_rank[1]['name']}，次者{by_rank[2]['nickname']}"
         f"{by_rank[2]['name']}。"),
        "",
        (f"军师二人：{by_rank[3]['nickname']}{by_rank[3]['name']}、"
         f"{by_rank[4]['nickname']}{by_rank[4]['name']}，同掌机密。"),
        "",
        "## 二 · 五虎将",
        "",
        f"马军五虎将共 {conflict_tiger} 人，计有：" + "、".join(
            f"{r['nickname']}{r['name']}" for r in rows
            if r["role_group"] == "马军五虎将") + "。",
        "",
        "## 三 · 战功",
        "",
        (f"步军头领{wusong['nickname']}{wusong['name']}，累计战功 {conflict_merit} 次，"
         f"为步军之最。"),
        "",
        f"水军头领共 {conflict_navy} 人，专掌舟楫。",
        "",
        "## 四 · 招安",
        "",
        (f"招安之议起，众头领中主张受诏者以{by_rank[1]['name']}为首。"
         f"名录所载上山年份最早者为 {conflict_year} 年，"
         f"最晚者为 {max(r['join_year'] for r in rows)} 年。"),
        "",
    ]
    _write(out / "docs" / "zhaoan.md", "\n".join(md))

    # PDF（文本型，非扫描件；不引入 OCR 依赖）
    from reportlab import rl_config
    rl_config.invariant = 1   # 固定 /CreationDate 与文档 ID，使 PDF 字节可复现

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    body = ParagraphStyle("body", parent=styles["Normal"],
                          fontName="STSong-Light", fontSize=11, leading=18)
    head = ParagraphStyle("head", parent=styles["Heading2"],
                          fontName="STSong-Light", fontSize=14, leading=22)

    pdf_path = out / "docs" / "zhaoan.pdf"
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                            title="梁山招安始末 · 改编节选")
    flow = []
    for line in md:
        if not line.strip() or line.startswith(">"):
            continue
        if line.startswith("#"):
            flow.append(Paragraph(line.lstrip("# ").strip(), head))
        else:
            flow.append(Paragraph(line, body))
        flow.append(Spacer(1, 4))
    doc.build(flow)

    _write(out / "perturbation_manifest.json",
           json.dumps(manifest, ensure_ascii=False, indent=2) + "\n")

    return {"book": "shuihu", "rows": len(rows),
            "xlsx_sha256": _sha256(xlsx_path), "pdf_sha256": _sha256(pdf_path),
            "manifest": manifest}


# ---------------------------------------------------------------- 语料 README

README_TMPL = """# {title} · 语料包

> 由 `evals/scripts/build_corpus.py` 生成。**不要手工编辑本目录下的产物**——改基础数据后重跑脚本。

## 数据形态

{formats}

## 字段来源分级

| 级别 | 含义 | 本包中的字段 |
| --- | --- | --- |
| canonical | 原著可考，存在于 LLM 预训练数据中 | {canonical} |
| curated | 人工整理，口径为本项目自定义，**待人工复核后冻结 v1** | {curated} |
| synthetic | 本项目虚构，固定种子生成，世界知识无法作答 | {synthetic} |

**为什么要混合 canonical 与 synthetic**：canonical 字段被扰动后，用于测量"系统是否真的查了数据"
（世界知识会给出与数据冲突的答案）；synthetic 字段任何模型都不可能知道，用于测量"纯检索与聚合"能力。
两类互补，报告中分别统计。

## 判分纪律

**ground truth 一律以本目录的实际数据 + `perturbation_manifest.json` 为准，不以原著为准。**
基础数据中若存在与原著的个别出入，不影响判分有效性；但会影响裸 LLM 基线的解释力，
故 curated 字段标记为待复核。

## 扰动

共 {n_pert} 条声明式扰动，逐条记录 before/after 与理由，见 `perturbation_manifest.json`。
`validate_corpus.py` 会逐条回验扰动是否已生效。

## 受控跨源冲突

{conflicts}

这些冲突是**故意的**。正确行为是报告分歧并给出两处出处，而不是擅自选一个。
"""


def write_readme(book: str, meta: dict) -> None:
    cfg = {
        "xiyouji": {
            "title": "西游记 · 八十一难",
            "formats": ("- `sql/xiyouji.db` —— SQLite，`nan`（81 行）与 `characters`（25 行）两表\n"
                     "- `csv/nan.csv`、`csv/characters.csv` —— 同一数据的 CSV 形态\n"
                     "- `docs/xiyouji_notes.md` —— 改编节选（新创作文本）"),
            "canonical": "`nan_no` / `nan_name`",
            "curated": "`category` / `location` / `opponent` / `helper` / `chapter_hint`",
            "synthetic": "`duration_days` / `difficulty_score`"},
        "shuihu": {
            "title": "水浒传 · 一百单八将",
            "formats": ("- `xlsx/liangshan.xlsx` —— Excel，`头领名录`（108 行）与 `职司汇总` 两 sheet\n"
                     "  - **故意保留三处真实世界的脏结构**：首行合并单元格标题（真实表头在第 2 行）、"
                     "中文表头带前后空格、`上山年份` 列存为文本\n"
                     "- `docs/zhaoan.md` / `docs/zhaoan.pdf` —— 改编节选（文本型 PDF，非扫描件）"),
            "canonical": "`rank` / `star_name` / `nickname` / `name`",
            "curated": "`role_group`",
            "synthetic": "`home_town` / `join_year` / `merit_count` / `outcome`"},
    }[book]
    conflicts = "\n".join(
        f"- **{c['id']}**：{c['fact']} —— 文档侧记为 `{c['doc_value']}`，与表格侧不一致。"
        f"期望行为：{c['expected_behavior']}"
        for c in CONFLICTS.get(book, [])) or "（无）"
    _write(NOVELS / book / "README.md",
           README_TMPL.format(n_pert=meta["manifest"]["perturbation_count"],
                              conflicts=conflicts, **cfg))


def main() -> None:
    results = [build_xiyouji(), build_shuihu()]
    for meta in results:
        write_readme(meta["book"], meta)
        print(f"[build] {meta['book']}: {meta['rows']} 行，"
              f"{meta['manifest']['perturbation_count']} 条扰动，"
              f"{len(meta['manifest']['conflicts'])} 处受控冲突")
    print("[build] 完成 → evals/novels/")


if __name__ == "__main__":
    main()
