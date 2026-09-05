"""语料与题集校验器（S0 能力门）。

一个基准的价值取决于它是否值得信任。本脚本逐条回验：

1. **四形态一致性** —— SQLite / CSV / XLSX 承载的是同一份数据
2. **扰动可复算** —— manifest 里每条扰动都能在语料中定位到，且 after 值与数据一致
3. **受控冲突符合声明** —— 文档值确实与表值不同，且差异符合声明
4. **ground truth 自校验** —— 每道模板题的 source_sql 重跑一次，结果必须等于存档答案
5. **题集配额** —— 各题型数量符合 BENCHMARK.md 的矩阵
6. **管线确定性** —— 由基础数据重新推导，结果必须与已发布产物逐行一致
7. **风险数据自洽** —— 难度档与信噪比一致、无幽灵模式、植入模式在数据中真实可观测

任何一条失败即非零退出。CI 直接用它当门禁。

用法::

    python evals/scripts/validate_corpus.py
"""

from __future__ import annotations

import csv
import json
import sqlite3
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))

from data import shuihu as shuihu_data
from data import xiyouji as xiyouji_data
from gen_questions import load_shuihu, load_xiyouji
from gen_risk_data import (
    DISTRACTORS,
    PLANTED,
    check_declarations,
    check_plantability,
)
from perturb import CONFLICTS, apply_perturbations

ROOT = Path(__file__).resolve().parents[1]
NOVELS = ROOT / "novels"
QUESTIONS = ROOT / "questions"
RISK = ROOT / "risk"
KNOWLEDGE = ROOT / "knowledge"

FAILURES: list[str] = []
CHECKS = 0


def check(cond: bool, label: str, detail: str = "") -> None:
    global CHECKS
    CHECKS += 1
    if not cond:
        FAILURES.append(f"{label}{(' — ' + detail) if detail else ''}")


# ---------------------------------------------------------------- 1/6 一致性与确定性

def check_xiyouji() -> None:
    con = sqlite3.connect(NOVELS / "xiyouji" / "sql" / "xiyouji.db")
    con.row_factory = sqlite3.Row
    db_rows = [dict(r) for r in con.execute("SELECT * FROM nan ORDER BY nan_no")]
    con.close()

    with (NOVELS / "xiyouji" / "csv" / "nan.csv").open(encoding="utf-8") as f:
        csv_rows = list(csv.DictReader(f))

    check(len(db_rows) == 81, "西游记 SQLite 行数应为 81", f"实际 {len(db_rows)}")
    check(len(csv_rows) == 81, "西游记 CSV 行数应为 81", f"实际 {len(csv_rows)}")

    # 四形态一致性：SQLite vs CSV（CSV 全为字符串，逐字段字符串化后比对）
    mismatch = []
    for d, c in zip(db_rows, csv_rows, strict=True):
        for k, v in d.items():
            cv = c[k]
            dv = "" if v is None else str(v)
            if dv != cv:
                mismatch.append(f"nan_no={d['nan_no']} 字段 {k}: db={dv!r} csv={cv!r}")
    check(not mismatch, "西游记 SQLite 与 CSV 不一致", "; ".join(mismatch[:3]))

    # 管线确定性：由基础数据重新推导，应与已发布产物完全一致
    rebuilt, manifest = apply_perturbations("xiyouji", xiyouji_data.rows())
    check(rebuilt == db_rows, "西游记语料无法由基础数据 + 扰动声明复现",
          "已发布产物与重新推导结果不一致")

    # 扰动可复算
    shipped = json.loads((NOVELS / "xiyouji" / "perturbation_manifest.json").read_text("utf-8"))
    check(shipped["perturbation_count"] == len(manifest["perturbations"]),
          "西游记扰动条数与声明不符")
    by_no = {r["nan_no"]: r for r in db_rows}
    for p in shipped["perturbations"]:
        for key, after in zip(p["keys"], p["after"], strict=True):
            check(by_no[key][p["field"]] == after,
                  f"扰动 {p['id']} 未在语料中生效",
                  f"nan_no={key} {p['field']} 期望 {after!r} 实际 {by_no[key][p['field']]!r}")
        check(p["before"] != p["after"], f"扰动 {p['id']} 的 before 与 after 相同（无效扰动）")

    # 受控冲突：文档值必须与表值不同
    doc = (NOVELS / "xiyouji" / "docs" / "xiyouji_notes.md").read_text("utf-8")
    check(str(CONFLICTS["xiyouji"][0]["doc_value"]) in doc,
          "XY-C01 的文档侧取值未出现在 Markdown 中")
    check(by_no[36]["location"] != CONFLICTS["xiyouji"][0]["doc_value"],
          "XY-C01 未构成冲突：表值与文档值相同")
    check(by_no[81]["opponent"] != CONFLICTS["xiyouji"][3]["doc_value"],
          "XY-C04 未构成冲突：表值与文档值相同")


def check_shuihu() -> None:
    con = load_shuihu()
    rows = [dict(r) for r in con.execute("SELECT * FROM liangshan ORDER BY rank")]
    con.close()
    check(len(rows) == 108, "水浒 Excel 行数应为 108", f"实际 {len(rows)}")

    rebuilt, _ = apply_perturbations("shuihu", shuihu_data.rows())
    check(rebuilt == rows, "水浒语料无法由基础数据 + 扰动声明复现",
          "Excel 读回结果与重新推导不一致（注意年份列存为文本）")

    shipped = json.loads((NOVELS / "shuihu" / "perturbation_manifest.json").read_text("utf-8"))
    by_rank = {r["rank"]: r for r in rows}
    for p in shipped["perturbations"]:
        for key, after in zip(p["keys"], p["after"], strict=True):
            check(by_rank[key][p["field"]] == after,
                  f"扰动 {p['id']} 未在语料中生效",
                  f"rank={key} {p['field']} 期望 {after!r} 实际 {by_rank[key][p['field']]!r}")

    # Excel 的三处脏结构必须真的存在（否则测不到单元格溯源的难点）
    from openpyxl import load_workbook
    ws = load_workbook(NOVELS / "shuihu" / "xlsx" / "liangshan.xlsx")["头领名录"]
    check(len(ws.merged_cells.ranges) >= 1, "Excel 缺少合并单元格陷阱")
    header = [ws.cell(row=2, column=c).value for c in range(1, 11)]
    check(any(isinstance(h, str) and h != h.strip() for h in header),
          "Excel 表头缺少带空格的陷阱")
    check(isinstance(ws.cell(row=3, column=8).value, str),
          "Excel 的『上山年份』列应存为文本（数字存文本陷阱）")


# ---------------------------------------------------------------- 4/5 题集

def check_questions() -> None:
    expect = {"F1": 16, "F2": 8, "F3": 14, "F4": 12, "F5": 6, "F6": 4}
    loaders = {"xiyouji": load_xiyouji, "shuihu": load_shuihu}

    for book in ("xiyouji", "shuihu"):
        items: list[dict[str, Any]] = []
        for path in (QUESTIONS / f"{book}.jsonl", QUESTIONS / "manual" / f"{book}.jsonl"):
            check(path.exists(), f"题集文件缺失: {path.relative_to(ROOT)}")
            if not path.exists():
                continue
            for lineno, line in enumerate(path.read_text("utf-8").splitlines(), 1):
                if not line.strip():
                    continue
                try:
                    items.append(json.loads(line))
                except json.JSONDecodeError as e:
                    check(False, f"题集 {path.name} 第 {lineno} 行不是合法 JSON", str(e))

        counts: dict[str, int] = {}
        for it in items:
            counts[it["type"]] = counts.get(it["type"], 0) + 1
        check(counts == expect, f"{book} 题型配额不符", f"期望 {expect} 实际 {counts}")
        check(len(items) == 60, f"{book} 题数应为 60", f"实际 {len(items)}")

        ids = [it["id"] for it in items]
        check(len(ids) == len(set(ids)), f"{book} 题目 id 重复")

        # ground truth 自校验：source_sql 重跑必须等于存档答案
        con = loaders[book]()
        bad = []
        for it in items:
            if not it.get("source_sql"):
                continue
            cur = con.execute(it["source_sql"], tuple(it.get("source_params") or ()))
            fetched = cur.fetchall()
            if it["answer_kind"] == "set":
                got: Any = sorted(str(r[0]) for r in fetched)
            elif it["answer_kind"] == "number":
                got = round(float(fetched[0][0]), 4)
            else:
                got = fetched[0][0] if fetched else None
            if got != it["ground_truth"]:
                bad.append(f"{it['id']}: 存档 {it['ground_truth']!r} 重算 {got!r}")
        con.close()
        check(not bad, f"{book} 存在 ground truth 与 source_sql 重算结果不符的题",
              "; ".join(bad[:3]))

        # F6 必须是真正不可答的
        for it in items:
            if it["type"] == "F6":
                check(it["ground_truth"].get("answerable") is False,
                      f"{it['id']} 标为 F6 却可答")
                check(it["trace_required"] is False, f"{it['id']} 拒答题不应要求溯源")

        check_units(book, items)


def check_units(book: str, items: list[dict[str, Any]]) -> None:
    """统计单位契约：确认性检验的分析单位必须在资产里冻结，不能由下游 loader 猜。

    这里挡住三类会直接污染主结论的错误：
    1. 缺 `template_id` → contamination diagnostic 的 cluster key 落空；
    2. F5 的三个 turn 被当成三行独立样本 → McNemar 伪重复；
    3. 残留旧字段 `perturbed` / `baseline_bucket` → 前者会被误读成数据集分区，
       后者曾指向一个并不存在的 run_baseline.py。
    """
    required = ("template_id", "independence_unit_id", "unit_kind")
    for key in required:
        missing = [it["id"] for it in items if not it.get(key)]
        check(not missing, f"{book} 有题目缺少 {key}", "; ".join(missing[:3]))

    legacy = [it["id"] for it in items
              if "perturbed" in it or "baseline_bucket" in it]
    check(not legacy, f"{book} 残留已废弃字段 perturbed/baseline_bucket", "; ".join(legacy[:3]))

    bad_kind = [it["id"] for it in items if it.get("unit_kind") not in ("single", "chain")]
    check(not bad_kind, f"{book} unit_kind 取值非法", "; ".join(bad_kind[:3]))

    # single → unit 就是题本身；chain → unit 是 chain_id，且只有 F5 是 chain
    bad_single = [it["id"] for it in items
                  if it.get("unit_kind") == "single" and it.get("independence_unit_id") != it["id"]]
    check(not bad_single, f"{book} 单轮题的 independence_unit_id 应等于 id",
          "; ".join(bad_single[:3]))
    bad_chain = [it["id"] for it in items
                 if it.get("unit_kind") == "chain"
                 and it.get("independence_unit_id") != it.get("chain_id")]
    check(not bad_chain, f"{book} 多轮题的 independence_unit_id 应等于 chain_id",
          "; ".join(bad_chain[:3]))
    mismatched = [it["id"] for it in items
                  if (it.get("unit_kind") == "chain") != (it["type"] == "F5")]
    check(not mismatched, f"{book} 只有 F5 允许 unit_kind=chain", "; ".join(mismatched[:3]))

    # 每部：60 题 → 54 个单轮 unit + 2 条 chain = 56 个 independence unit
    units = {it["independence_unit_id"] for it in items}
    check(len(units) == 56, f"{book} independence unit 数应为 56", f"实际 {len(units)}")
    chains: dict[str, int] = {}
    for it in items:
        if it.get("unit_kind") == "chain":
            chains[it["independence_unit_id"]] = chains.get(it["independence_unit_id"], 0) + 1
    check(sorted(chains.values()) == [3, 3], f"{book} 应有 2 条各 3 轮的 chain", f"实际 {chains}")

    # 一个 template_id 只能属于一个题型，否则 cluster 会横跨异质题型。
    # 用 .get 兜底：check() 只记录失败不中断，缺字段时若直接下标会抛 KeyError，
    # 把其余上百项校验的报告一起吞掉。
    tpl_types: dict[str, set[str]] = {}
    for it in items:
        tpl_types.setdefault(it.get("template_id") or "<missing>", set()).add(it["type"])
    cross = [t for t, ts in tpl_types.items() if len(ts) > 1]
    check(not cross, f"{book} 存在跨题型复用的 template_id", "; ".join(cross[:3]))

    # F4 是唯一以被扰动字段为目标的题型；该标记是题目属性而非数据集分区
    bad_flag = [it["id"] for it in items
                if bool(it.get("targets_perturbed_field")) != (it["type"] == "F4")]
    check(not bad_flag, f"{book} targets_perturbed_field 只应在 F4 上为真",
          "; ".join(bad_flag[:3]))


# ---------------------------------------------------------------- 7 风险数据

def check_risk() -> None:
    check_declarations()
    check_plantability()

    manifest = json.loads((RISK / "planted_manifest.json").read_text("utf-8"))
    check(len(manifest["planted"]) == len(PLANTED), "植入清单条数与声明不符")
    check(len(manifest["distractors"]) == len(DISTRACTORS), "干扰项条数与声明不符")
    check(len(manifest["limitations"]) >= 4, "植入清单缺少局限声明")

    con = sqlite3.connect(RISK / "yunti.db")

    # 逐条抽验：植入模式必须在数据中真实可观测（方向正确、幅度达标）
    def series(sql: str, params: tuple = ()) -> list[float]:
        return [r[0] for r in con.execute(sql, params) if r[0] is not None]

    p1 = series("SELECT AVG(gross_margin_rate) FROM fact_revenue "
                "WHERE product_line='云梯-企业版' GROUP BY month ORDER BY month")
    check(p1[5] - p1[-1] > 0.05, "P1-001 毛利率下滑未在数据中体现",
          f"首末差 {p1[5] - p1[-1]:.4f}")

    p4 = series("SELECT dso_days FROM fact_receivable WHERE customer_id='C03' ORDER BY month")
    check(max(p4) / min(p4) > 2.0, "P4-001 账期跳升未在数据中体现")

    d3 = series("SELECT dso_days FROM fact_receivable WHERE customer_id='C08' ORDER BY month")
    check(max(d3) / min(d3) < 1.8, "干扰项 D-003 幅度过大，会被当成真异常",
          f"极值比 {max(d3) / min(d3):.2f}")

    p5 = series("SELECT cash_to_revenue FROM fact_cashflow "
                "WHERE product_line='云梯-标准版' ORDER BY month")
    check(p5[0] - p5[-1] > 0.02, "P5-002 现金流背离未在数据中体现")

    # 脏度必须真的存在（否则"合成数据太干净"的缓解措施等于没做）
    nulls = con.execute("SELECT COUNT(*) FROM fact_receivable WHERE overdue IS NULL").fetchone()[0]
    check(nulls > 0, "应收表未注入缺失值")
    dup = con.execute("SELECT COUNT(*) FROM (SELECT month,product_line,region,COUNT(*) c "
                      "FROM fact_revenue GROUP BY 1,2,3 HAVING c>1)").fetchone()[0]
    check(dup > 0, "营收表未注入重复记录")
    con.close()


# ---------------------------------------------------------------- 8 S5 knowledge corpus

# ADR-013 决策 2：evals/knowledge/ 是评测先行的冻结语料（specs/s5 §6），必须纳入
# validator 口径。与 120 题的差异：ground truth 是内容值/行为标签而非 SQL 重算——
# 可复算回退到两个层面：结构一致性 + 行为参数与 ground truth 语义的可复算。

KNOWLEDGE_SUITES = {
    "doc_table_v1.jsonl": "knowledge-doc-v1",
    "code_github_v1.jsonl": "knowledge-code-github-v1",
    "cross_source_v1.jsonl": "knowledge-cross-source-v1",
    "acl_freshness_v1.jsonl": "knowledge-acl-freshness-v1",
}

KNOWLEDGE_MIN_BLIND_HOLDOUT = {
    "knowledge-doc-v1": 3,
    "knowledge-code-github-v1": 2,
    "knowledge-cross-source-v1": 2,
    "knowledge-acl-freshness-v1": 2,
}

KNOWLEDGE_CLASSIFICATION_ORDER = {
    "PUBLIC": 0,
    "INTERNAL": 1,
    "CONFIDENTIAL": 2,
    "RESTRICTED": 3,
}

KNOWLEDGE_METAMORPHIC_VARIANTS = {"rename", "move", "update", "revoke"}


def check_knowledge() -> None:
    all_ids: dict[str, str] = {}
    for filename, suite in KNOWLEDGE_SUITES.items():
        path = KNOWLEDGE / filename
        check(path.exists(), f"knowledge 语料文件缺失: {path.relative_to(ROOT)}")
        if not path.exists():
            continue
        items: list[dict[str, Any]] = []
        for lineno, line in enumerate(path.read_text("utf-8").splitlines(), 1):
            if not line.strip():
                continue
            try:
                items.append(json.loads(line))
            except json.JSONDecodeError as e:
                check(False, f"knowledge 语料 {filename} 第 {lineno} 行不是合法 JSON", str(e))
        check(len(items) > 0, f"knowledge 语料 {filename} 为空")

        ids = [it["id"] for it in items]
        check(len(ids) == len(set(ids)), f"knowledge 语料 {filename} 题目 id 重复")
        base_ids = {it["id"] for it in items if not it.get("metamorphic_variant")}
        blind = [it for it in items if it.get("blind_holdout")]
        check(
            len(blind) >= KNOWLEDGE_MIN_BLIND_HOLDOUT[suite],
            f"knowledge 语料 {filename} blind holdout 数不足",
            f"实际 {len(blind)}",
        )

        for it in items:
            loc = f"{filename}:{it['id']}"
            check(it.get("suite") == suite, f"{loc} suite 字段与文件不符",
                  f"{it.get('suite')!r} != {suite!r}")
            check(it.get("unit_kind") in ("single", "chain"),
                  f"{loc} unit_kind 取值非法", str(it.get("unit_kind")))
            if it.get("unit_kind") == "single":
                check(it.get("independence_unit_id") == it["id"],
                      f"{loc} 单轮单位的 independence_unit_id 应等于 id")

            check(it.get("answer_kind") in ("scalar", "number", "set"),
                  f"{loc} answer_kind 取值非法", str(it.get("answer_kind")))
            mode = (it.get("scoring") or {}).get("mode")
            check(mode in ("exact", "numeric", "contains", "set"),
                  f"{loc} scoring.mode 取值非法", str(mode))
            gt = it.get("ground_truth")
            if it.get("answer_kind") == "set":
                check(isinstance(gt, list) and all(isinstance(x, str) for x in gt),
                      f"{loc} set 答案应为字符串列表", repr(gt))
            else:
                check(isinstance(gt, str), f"{loc} {it.get('answer_kind')} 答案应为字符串",
                      repr(gt))
            check(it.get("trace_required") is True, f"{loc} knowledge 语料必须要求溯源")

            variant = it.get("metamorphic_variant")
            if variant is not None:
                check(variant in KNOWLEDGE_METAMORPHIC_VARIANTS,
                      f"{loc} metamorphic 变体取值非法", str(variant))
                base_id = it.get("metamorphic_base_id")
                base = next((b for b in items if b["id"] == base_id), None)
                check(base is not None and base_id in base_ids,
                      f"{loc} metamorphic_base_id 必须指向本文件非变体条目", repr(base_id))
                if base is not None:
                    check(it.get("answer_kind") == base.get("answer_kind"),
                          f"{loc} answer_kind 与 base 不一致")
            else:
                check(it.get("metamorphic_base_id") is None,
                      f"{loc} 非变体条目携带 metamorphic_base_id")

            if variant != "revoke":
                check(len(it.get("expected_locators") or []) > 0,
                      f"{loc} 非 revoke 条目必须有 expected_locators")
            for i, locator in enumerate(it.get("expected_locators") or []):
                check(isinstance(locator.get("connector"), str) and locator["connector"],
                      f"{loc} locator[{i}] 缺少 connector")
                check(isinstance(locator.get("uri"), str) and locator["uri"],
                      f"{loc} locator[{i}] 缺少 uri")

            # 可复算：声明的行为参数必须能推出声明的 ground truth 语义
            qt = it.get("query_type")
            if qt == "freshness_stale":
                check(it.get("freshness_age_days", 0) > it.get("aging_threshold_days", 0),
                      f"{loc} stale 场景参数推不出 aged")
            elif qt == "freshness_fresh":
                check(it.get("freshness_age_days", 1) < it.get("aging_threshold_days", 0),
                      f"{loc} fresh 场景参数推不出 fresh")
            elif qt == "acl_pre_filter":
                clearance_value = it.get("acl_clearance")
                target_value = it.get("target_classification")
                clearance = (
                    KNOWLEDGE_CLASSIFICATION_ORDER.get(clearance_value)
                    if isinstance(clearance_value, str)
                    else None
                )
                target = (
                    KNOWLEDGE_CLASSIFICATION_ORDER.get(target_value)
                    if isinstance(target_value, str)
                    else None
                )
                check(clearance is not None and target is not None and clearance < target,
                      f"{loc} pre-filter 场景 clearance 必须低于目标分类")
            elif qt == "cross_org_query":
                check(it.get("query_org") != it.get("target_org"),
                      f"{loc} 跨组织场景 query_org 与 target_org 相同")
            if it.get("revoked_at_query_time") is True:
                check(qt == "acl_hydration_recheck",
                      f"{loc} revoked_at_query_time 只应出现在 acl_hydration_recheck")

            prior = all_ids.get(it["id"])
            check(prior is None, "knowledge 语料 id 跨文件重复",
                  f"{it['id']} 同时出现在 {prior} 与 {filename}")
            if prior is None:
                all_ids[it["id"]] = filename

    # 与 legacy 120 题无 id 重叠（S5-T8 集成测试同约束；validator 独立复核）
    legacy_ids = set()
    for path in QUESTIONS.glob("*.jsonl"):
        for line in path.read_text("utf-8").splitlines():
            if line.strip():
                legacy_ids.add(json.loads(line)["id"])
    overlap = set(all_ids) & legacy_ids
    check(not overlap, "knowledge 语料 id 与 legacy 题集重叠", "; ".join(sorted(overlap)[:3]))


def check_checksums() -> None:
    """篡改检测：已发布产物必须与 CHECKSUMS.sha256 一致。

    捕获两类事故：有人手改了生成产物（数据即代码的前提被破坏），
    或改了生成脚本却忘了重跑（题集与语料脱节）。
    """
    import hashlib

    sums = ROOT / "CHECKSUMS.sha256"
    if not sums.exists():
        check(False, "缺少 evals/CHECKSUMS.sha256", "先跑 make checksums")
        return

    recorded = {}
    for line in sums.read_text("utf-8").splitlines():
        if line.strip():
            digest, name = line.split(None, 1)
            recorded[name.strip()] = digest

    current = {}
    for d in ("novels", "questions", "risk", "knowledge"):
        for p in sorted((ROOT / d).rglob("*")):
            if p.is_file():
                rel = str(p.relative_to(ROOT.parent))
                current[rel] = hashlib.sha256(p.read_bytes()).hexdigest()

    missing = sorted(set(recorded) - set(current))
    added = sorted(set(current) - set(recorded))
    changed = sorted(k for k in set(recorded) & set(current) if recorded[k] != current[k])
    check(not missing, "校验和记录的产物已丢失", "; ".join(missing[:3]))
    check(not added, "存在未登记的新产物（请重跑 make checksums）", "; ".join(added[:3]))
    check(not changed, "产物与校验和不符（被手工修改或脚本已变更未重跑）",
          "; ".join(changed[:3]))


def main() -> None:
    check_xiyouji()
    check_shuihu()
    check_questions()
    check_risk()
    check_knowledge()
    check_checksums()

    if FAILURES:
        print(f"[validate] {len(FAILURES)}/{CHECKS} 项校验失败：")
        for f in FAILURES:
            print(f"  ✗ {f}")
        raise SystemExit(1)
    print(f"[validate] {CHECKS} 项校验全部通过")


if __name__ == "__main__":
    main()
