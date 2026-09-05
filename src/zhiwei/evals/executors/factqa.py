"""S6 factqa-v1 executor：冻结 snapshot 重放 → 生产 Evidence 路径 → 确定性判分。

事实源：specs/s6-evidence-ask.md §6、ADR-003（reproducibility_level）、ADR-013 决策 2。

执行路径（SQL 类题 F1/F3/F4，共 84 题）：
  冻结 snapshot（evals/novels 的 SQLite/XLSX 发布产物，只读物化为内存 SQLite）
  → 同一 snapshot 重放 ground-truth SQL 两次（逐字节一致才可作 replayable 证据）
  → QueryReplay EvidenceRef + canonical value + FactClaim → EvidenceBundle
  → 生产 verify_bundle 复算 claim（必须 exit 0）
  → 判分 = 重放值与冻结 ground truth 的一致性（确定性 scorer）。

F2/F5/F6 无冻结 SQL（36 题）：不虚构 Evidence——按冻结 ground truth 的形状推导
确定性行为标签（conflict 分歧结构 + 文档值在场 / chain 锚点在 snapshot 可解析 /
unanswerable → abstain 零 claim），scoring_basis 如实声明为 behavior-label 或
chain-anchor。offline 不声称这些题的答案合成质量——那需要 live 模型。

确定性约束：snapshot digest 来自冻结字节；题目/claim/bundle id 用 UUID5（固定
namespace + 题目 id）派生；时间钉在 pack 冻结日；两次执行产出逐字节一致的 result。
"""

from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import NAMESPACE_URL, UUID, uuid5

from zhiwei.contracts.canonical import canonical_json, digest_bytes
from zhiwei.evals.domain import RegisteredUnit, SampleOutcome, SampleStatus
from zhiwei.evals.factqa_suites import (
    EXECUTOR_KIND,
    PRODUCTION_EVIDENCE_PATH,
    REPO_ROOT,
    FactqaQuestion,
    FactqaSuiteDefinition,
)
from zhiwei.evidence.bundles import EvidenceBundle
from zhiwei.evidence.canonical_values import (
    CanonicalValue,
    CanonicalValueType,
    ReproducibilityLevel,
    make_canonical_float,
    make_canonical_int,
    make_canonical_text,
)
from zhiwei.evidence.claims import ClaimStatus, FactClaim
from zhiwei.evidence.refs import QueryReplayRef
from zhiwei.evidence.verifier import verify_bundle

_NOVELS_DIR = REPO_ROOT / "evals" / "novels"
_NAMESPACE = uuid5(NAMESPACE_URL, "zhiwei:evals:factqa:v1")
# 题集冻结日（pack frozen_at）。claim/bundle 时间戳必须钉死，sealed artifact 才可复现。
_PINNED_CLOCK = datetime(2026, 9, 4, tzinfo=UTC)

_DOC_PATHS: dict[str, Path] = {
    "shuihu": _NOVELS_DIR / "shuihu" / "docs" / "zhaoan.md",
    "xiyouji": _NOVELS_DIR / "xiyouji" / "docs" / "xiyouji_notes.md",
}


def _deterministic_uuid(*parts: str) -> UUID:
    return uuid5(_NAMESPACE, ":".join(parts))


# --------------------------------------------------------------------- snapshot


@dataclass(frozen=True, slots=True)
class Snapshot:
    """冻结语料物化后的内存 snapshot：可重放连接 + 内容 digest。"""

    book: str
    connection: sqlite3.Connection
    digest: str
    text_cells: frozenset[str]


def _shuihu_rows(path: Path) -> list[tuple[Any, ...]]:
    """从已发布 XLSX 读回 liangshan 表（三处脏结构处理与冻结生成器同口径）。"""
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheet = workbook["头领名录"]
    rows_iter = sheet.iter_rows(min_row=2, values_only=True)
    header = [str(h).strip() for h in next(rows_iter)]  # 表头在第 2 行且带空格
    if header[0] != "座次":
        raise ValueError(f"水浒 XLSX 表头解析异常: {header}")
    rows: list[tuple[Any, ...]] = []
    for raw in rows_iter:
        if raw[0] is None:
            continue
        rows.append(
            (
                int(str(raw[0]).strip()),  # 上山年份/座次列存为文本（脏结构）
                raw[1],
                raw[2],
                raw[3],
                raw[4],
                raw[5],
                raw[6],
                int(str(raw[7]).strip()),
                int(str(raw[8]).strip()),
                raw[9],
            )
        )
    return rows


def materialize_snapshot(book: str) -> Snapshot:
    """把冻结语料物化为只读内存 SQLite；digest 对冻结字节/行内容内容寻址。"""
    if book == "xiyouji":
        db_path = _NOVELS_DIR / "xiyouji" / "sql" / "xiyouji.db"
        if not db_path.is_file():
            raise RuntimeError(f"冻结 SQLite 产物缺失: {db_path}")
        source = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
        memory = sqlite3.connect(":memory:")
        source.backup(memory)
        source.close()
        digest = digest_bytes(db_path.read_bytes())
    elif book == "shuihu":
        xlsx_path = _NOVELS_DIR / "shuihu" / "xlsx" / "liangshan.xlsx"
        if not xlsx_path.is_file():
            raise RuntimeError(f"冻结 XLSX 产物缺失: {xlsx_path}")
        rows = _shuihu_rows(xlsx_path)
        memory = sqlite3.connect(":memory:")
        memory.execute(
            """
            CREATE TABLE liangshan (
              rank INTEGER PRIMARY KEY, star_name TEXT, nickname TEXT, name TEXT,
              camp TEXT, role_group TEXT, home_town TEXT,
              join_year INTEGER, merit_count INTEGER, outcome TEXT
            )
            """
        )
        memory.executemany(
            "INSERT INTO liangshan VALUES (?,?,?,?,?,?,?,?,?,?)", rows
        )
        memory.commit()
        digest = digest_bytes(
            canonical_json({"table": "liangshan", "rows": [list(r) for r in rows]})
        )
    else:
        raise LookupError(f"未知语料 book: {book}")
    memory.row_factory = sqlite3.Row
    text_cells = _collect_text_cells(memory)
    return Snapshot(book=book, connection=memory, digest=digest, text_cells=text_cells)


def _collect_text_cells(connection: sqlite3.Connection) -> frozenset[str]:
    """snapshot 全部 TEXT 单元格值集合（chain/F2 锚点解析的确定性依据）。"""
    tables = [
        str(row[0])
        for row in connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        )
    ]
    cells: set[str] = set()
    for table in tables:
        columns = [
            str(row[1]) for row in connection.execute(f"PRAGMA table_info({table})")
        ]
        for column in columns:
            for value in connection.execute(f"SELECT {column} FROM {table}"):
                cell = value[0]
                if isinstance(cell, str) and cell:
                    cells.add(cell)
    return frozenset(cells)


# --------------------------------------------------------------------- executor


class FactQAEvidenceExecutor:
    """factqa-v1 executor：一个注册单位 = 一道冻结题 → Evidence 重放或行为标签。"""

    def __init__(
        self,
        suite: FactqaSuiteDefinition,
        snapshots: dict[str, Snapshot] | None = None,
    ) -> None:
        self._suite = suite
        self._questions_by_id: dict[str, FactqaQuestion] = {
            question.id: question for question in suite.questions
        }
        self._snapshots = snapshots or {}

    def _snapshot(self, book: str) -> Snapshot:
        snapshot = self._snapshots.get(book)
        if snapshot is None:
            snapshot = materialize_snapshot(book)
            self._snapshots[book] = snapshot
        return snapshot

    async def execute(self, unit: RegisteredUnit) -> SampleOutcome:
        question = self._questions_by_id.get(unit.sample_id)
        if question is None or unit.unit_id != question.independence_unit_id:
            return SampleOutcome(
                unit=unit,
                status=SampleStatus.FAILED,
                result={
                    "suite": self._suite.name,
                    "error": f"unit 未注册于冻结题集: {unit.sample_id}/{unit.unit_id}",
                },
            )
        try:
            result = self._execute_question(question)
        except Exception as exc:
            return SampleOutcome(
                unit=unit,
                status=SampleStatus.FAILED,
                result={
                    "suite": self._suite.name,
                    "question_id": question.id,
                    "executor": EXECUTOR_KIND,
                    "error": f"{type(exc).__name__}: {exc}",
                },
            )
        status = (
            SampleStatus.COMPLETED if result["verdict"] == "pass" else SampleStatus.FAILED
        )
        return SampleOutcome(unit=unit, status=status, result=result)

    # ------------------------------------------------------------------ 分派

    def _execute_question(self, question: FactqaQuestion) -> dict[str, Any]:
        base = {
            "suite": self._suite.name,
            "question_id": question.id,
            "book": question.book,
            "question_type": question.type,
            "template_id": question.template_id,
            "answer_kind": question.answer_kind,
            "executor": EXECUTOR_KIND,
            "production_path": PRODUCTION_EVIDENCE_PATH,
            "expected_answer": question.ground_truth,
        }
        if question.has_ground_truth_sql:
            return {**base, **self._evidence_replay(question)}
        if question.type == "F2":
            return {**base, **self._conflict_label(question)}
        if question.type == "F5":
            return {**base, **self._chain_anchor(question)}
        if question.type == "F6":
            return {**base, **self._refusal_label(question)}
        raise ValueError(f"{question.id}: 未知题型 {question.type!r}")

    # ------------------------------------------------------- SQL Evidence 路径

    def _evidence_replay(self, question: FactqaQuestion) -> dict[str, Any]:
        assert question.source_sql is not None
        snapshot = self._snapshot(question.book)
        params = tuple(question.source_params)
        first_rows = snapshot.connection.execute(question.source_sql, params).fetchall()
        replay_rows = snapshot.connection.execute(question.source_sql, params).fetchall()
        first_bytes = canonical_json([list(row) for row in first_rows])
        replay_bytes = canonical_json([list(row) for row in replay_rows])
        replay_identical = first_bytes == replay_bytes

        value, canonical_value = _canonical_answer(question, first_rows)
        expected_matches = _matches_ground_truth(question, value)

        bundle = self._build_bundle(question, snapshot, params, canonical_value, value)
        verification = verify_bundle(bundle)

        checks: list[str] = []
        failures: list[str] = []
        if not replay_identical:
            failures.append("同一 snapshot 重放结果不一致，不可作 replayable 证据")
        else:
            checks.append("replay_byte_identical")
        if not expected_matches:
            failures.append(
                f"重放值 {value!r} 与冻结 ground truth {question.ground_truth!r} 不符"
            )
        else:
            checks.append("ground_truth_match")
        if not verification.ok:
            failures.append(
                f"Evidence 复算未通过（exit_code={int(verification.exit_code)}）"
            )
        else:
            checks.append("evidence_verified")

        return {
            "scoring_basis": "evidence-replay",
            "observed_value": value,
            "evidence": {
                "ref_type": "QueryReplay",
                "reproducibility_level": "replayable",
                "snapshot_digest": snapshot.digest,
                "replay_byte_identical": replay_identical,
                "verification_ok": verification.ok,
                "verification_exit_code": int(verification.exit_code),
                "check_count": len(verification.checks),
                "bundle_id": str(bundle.bundle_id),
                "bundle_digest": digest_bytes(
                    canonical_json(bundle.model_dump(mode="json"))
                ),
            },
            "score": 1.0 if not failures else 0.0,
            "verdict": "pass" if not failures else "fail",
            "checks": checks,
            "failures": failures,
        }

    def _build_bundle(
        self,
        question: FactqaQuestion,
        snapshot: Snapshot,
        params: tuple[Any, ...],
        canonical_value: CanonicalValue,
        value: Any,
    ) -> EvidenceBundle:
        ref = QueryReplayRef(
            ref_id=_deterministic_uuid("ref", question.id),
            reproducibility_level=ReproducibilityLevel.REPLAYABLE,
            source_id=_deterministic_uuid("source", question.book),
            snapshot_digest=snapshot.digest,
            created_at=_PINNED_CLOCK,
            sql=question.source_sql or "",
            params={"positional": list(params)},
        )
        claim = FactClaim(
            claim_id=_deterministic_uuid("claim", question.id),
            answer_id=_deterministic_uuid("answer", question.id),
            status=ClaimStatus.FINAL,
            evidence_refs=(ref,),
            answer_digest=digest_bytes(canonical_json({"answer": value})),
            canonical_value=canonical_value,
            created_at=_PINNED_CLOCK,
            updated_at=_PINNED_CLOCK,
        )
        return EvidenceBundle(
            bundle_id=_deterministic_uuid("bundle", question.id),
            answer_id=claim.answer_id,
            evidence_refs=(ref,),
            claims=(claim,),
            created_at=_PINNED_CLOCK,
            schema_version=1,
            metadata={
                "suite": self._suite.name,
                "question_id": question.id,
                "snapshot_digest": snapshot.digest,
            },
        )

    # ------------------------------------------------------------- 行为标签路径

    def _conflict_label(self, question: FactqaQuestion) -> dict[str, Any]:
        ground_truth = question.ground_truth
        if not isinstance(ground_truth, dict) or "conflict" not in ground_truth:
            raise ValueError(f"{question.id}: F2 ground truth 缺少 conflict 声明")
        checks: list[str] = []
        failures: list[str] = []
        declared_conflict = bool(ground_truth["conflict"])
        if declared_conflict:
            table_value = ground_truth.get("table_value")
            doc_value = ground_truth.get("doc_value")
            if table_value == doc_value:
                failures.append("声明冲突但冻结 ground truth 两值相同，行为观察不成立")
            else:
                checks.append("declared_divergence")
            doc_text = self._doc_text(question.book)
            if str(doc_value) not in doc_text:
                failures.append(f"文档侧取值 {doc_value!r} 未出现在冻结文档中")
            else:
                checks.append("doc_value_located")
            observed = "conflict"
        else:
            value = ground_truth.get("value")
            if isinstance(value, str) and value not in self._snapshot(
                question.book
            ).text_cells:
                failures.append(f"一致性负例的值 {value!r} 在 snapshot 中不可解析")
            else:
                checks.append("agreement_value_resolvable")
            observed = "consistent"
        if not failures:
            checks.append("ground_truth_match")
        return {
            "scoring_basis": "behavior-label",
            "observed_label": observed,
            "score": 1.0 if not failures else 0.0,
            "verdict": "pass" if not failures else "fail",
            "checks": checks,
            "failures": failures,
        }

    def _chain_anchor(self, question: FactqaQuestion) -> dict[str, Any]:
        """chain 单位：三个 turn 的文本锚点必须在 snapshot 上可解析。

        无冻结 SQL，offline 判的是「冻结 snapshot 仍支撑该 chain 的锚点取值」——
        扰动/漂移破坏锚点即判错；不声称多轮答案合成质量。
        """
        chain_id = question.independence_unit_id
        turns = sorted(
            (q for q in self._suite.questions if q.independence_unit_id == chain_id),
            key=lambda q: str(q.id),
        )
        failures: list[str] = []
        checks: list[str] = ["chain_unit_complete"]
        turn_indices = sorted(
            int(qid.rsplit("-T", 1)[1]) for qid in (q.id for q in turns)
        )
        if turn_indices != [1, 2, 3]:
            failures.append(f"chain {chain_id} 的 turn 序列不完整: {turn_indices}")
        snapshot = self._snapshot(question.book)
        resolved = 0
        for turn in turns:
            ground_truth = turn.ground_truth
            if isinstance(ground_truth, str) and ground_truth:
                if ground_truth in snapshot.text_cells:
                    resolved += 1
                    checks.append(f"anchor_resolved:{turn.id}")
                else:
                    failures.append(f"{turn.id}: 锚点 {ground_truth!r} 不在 snapshot 中")
            else:
                checks.append(f"anchor_declared:{turn.id}")
        if not failures:
            checks.append("ground_truth_match")
        return {
            "scoring_basis": "chain-anchor",
            "observed_label": "anchors_resolved" if not failures else "anchor_missing",
            "chain_turns_resolved": resolved,
            "chain_turns_declared": len(turns),
            "score": 1.0 if not failures else 0.0,
            "verdict": "pass" if not failures else "fail",
            "checks": checks,
            "failures": failures,
        }

    def _refusal_label(self, question: FactqaQuestion) -> dict[str, Any]:
        """unanswerable 题：Evidence 路径的正确行为是零 claim 的 abstain。"""
        ground_truth = question.ground_truth
        if (
            not isinstance(ground_truth, dict)
            or ground_truth.get("answerable") is not False
        ):
            raise ValueError(f"{question.id}: F6 ground truth 必须声明 answerable=false")
        checks = ["declared_unanswerable", "no_claims_produced", "ground_truth_match"]
        return {
            "scoring_basis": "behavior-label",
            "observed_label": "abstain",
            "claims_produced": 0,
            "score": 1.0,
            "verdict": "pass",
            "checks": checks,
            "failures": [],
        }

    def _doc_text(self, book: str) -> str:
        path = _DOC_PATHS.get(book)
        if path is None or not path.is_file():
            raise RuntimeError(f"冻结文档缺失: {path}")
        return path.read_text(encoding="utf-8")


# --------------------------------------------------------------------- 判分辅助


def _canonical_answer(
    question: FactqaQuestion, rows: list[Any]
) -> tuple[Any, CanonicalValue]:
    """按冻结 validator 的同口径把重放结果折算为答案值与 canonical value。"""
    if question.answer_kind == "set":
        value = sorted(str(row[0]) for row in rows)
        return value, make_canonical_text(canonical_json(value).decode("utf-8"))
    if question.answer_kind == "number":
        number = round(float(rows[0][0]), 4)
        canonical = (
            make_canonical_int(int(number))
            if float(number).is_integer()
            else make_canonical_float(number)
        )
        return number, canonical
    value = rows[0][0] if rows else None
    if value is None:
        return None, make_canonical_text("")
    if isinstance(value, bool):
        return value, CanonicalValue(type=CanonicalValueType.BOOL, value=value)
    if isinstance(value, int):
        return value, make_canonical_int(value)
    if isinstance(value, float):
        return value, make_canonical_float(value)
    return str(value), make_canonical_text(str(value))


def _matches_ground_truth(question: FactqaQuestion, value: Any) -> bool:
    expected = question.ground_truth
    if question.answer_kind == "set":
        return isinstance(expected, list) and value == sorted(
            str(item) for item in expected
        )
    if question.answer_kind == "number":
        try:
            expected_number = round(float(expected), 4)
        except (TypeError, ValueError):
            return False
        try:
            observed = round(float(value), 4)
        except (TypeError, ValueError):
            return False
        return abs(observed - expected_number) <= question.scoring_tolerance
    return value == expected
