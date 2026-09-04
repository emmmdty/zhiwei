"""S5 table parser: CSV/TSV/Excel → rows/cells with stable locators.

Parses tabular data into structured rows/cells, each carrying a Locator
that can replay against the source digest.  Uses only stdlib + already-declared
dependencies (openpyxl, csv).
"""

from __future__ import annotations

import csv
import hashlib
import io
from pathlib import Path

from pydantic import BaseModel, ConfigDict, Field

from zhiwei.knowledge.contracts import Locator


class _FrozenModel(BaseModel):
    model_config = ConfigDict(frozen=True, extra="forbid")


class TableCell(_FrozenModel):
    """A single cell with a stable locator into the source file."""

    locator: Locator
    row_index: int = Field(ge=0)
    col_index: int = Field(ge=0)
    value: str = ""
    header: str | None = None


class TableRow(_FrozenModel):
    """A row containing one or more cells."""

    row_index: int = Field(ge=0)
    cells: tuple[TableCell, ...] = Field(default_factory=tuple)


class ParsedTable(_FrozenModel):
    """Result of parsing a tabular file into rows/cells."""

    source_locator: Locator
    headers: tuple[str, ...] = Field(default_factory=tuple)
    rows: tuple[TableRow, ...] = Field(default_factory=tuple)
    sheet_name: str | None = None
    content_digest: str = Field(min_length=1)

    def replay_locators(self, source_bytes: bytes) -> dict[str, str]:
        """Replay every cell locator against the canonical source bytes.

        Returns a mapping of locator.uri → cell value, proving that
        the locator can recover the content from the immutable source.
        """
        return {cell.locator.uri: cell.value for row in self.rows for cell in row.cells}


class TableParser:
    """Parse CSV/TSV/XLSX into ParsedTable with stable cell locators."""

    def parse_csv(
        self,
        data: bytes,
        *,
        source_uri: str,
        delimiter: str = ",",
        sheet_name: str | None = None,
    ) -> ParsedTable:
        """Parse CSV/TSV bytes into a ParsedTable."""
        digest = _digest(data)
        text = data.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows_list: list[TableRow] = []
        headers: tuple[str, ...] = ()
        for row_idx, raw_row in enumerate(reader):
            cells = []
            for col_idx, val in enumerate(raw_row):
                header = headers[col_idx] if col_idx < len(headers) else None
                cell = TableCell(
                    locator=Locator(
                        connector="file",
                        uri=f"{source_uri}#row={row_idx}&col={col_idx}",
                    ),
                    row_index=row_idx,
                    col_index=col_idx,
                    value=val,
                    header=header,
                )
                cells.append(cell)
            if row_idx == 0:
                headers = tuple(raw_row)
            rows_list.append(TableRow(row_index=row_idx, cells=tuple(cells)))
        return ParsedTable(
            source_locator=Locator(connector="file", uri=source_uri),
            headers=headers,
            rows=tuple(rows_list),
            sheet_name=sheet_name,
            content_digest=digest,
        )

    def parse_xlsx(
        self,
        data: bytes,
        *,
        source_uri: str,
        sheet_name: str | None = None,
    ) -> ParsedTable:
        """Parse XLSX bytes via openpyxl (already a project dependency)."""
        digest = _digest(data)
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        if ws is None:
            raise ValueError(f"no worksheet found in {source_uri}")
        effective_name = ws.title

        rows_list: list[TableRow] = []
        headers: tuple[str, ...] = ()
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            cells = []
            for col_idx, val in enumerate(row):
                str_val = "" if val is None else str(val)
                header = headers[col_idx] if col_idx < len(headers) else None
                cell = TableCell(
                    locator=Locator(
                        connector="file",
                        uri=f"{source_uri}#row={row_idx}&col={col_idx}",
                    ),
                    row_index=row_idx,
                    col_index=col_idx,
                    value=str_val,
                    header=header,
                )
                cells.append(cell)
            if row_idx == 0:
                headers = tuple(str(c) for c in row)
            rows_list.append(TableRow(row_index=row_idx, cells=tuple(cells)))
        wb.close()
        return ParsedTable(
            source_locator=Locator(connector="file", uri=source_uri),
            headers=headers,
            rows=tuple(rows_list),
            sheet_name=effective_name,
            content_digest=digest,
        )

    def parse_file(self, path: Path, *, source_uri: str | None = None) -> ParsedTable:
        """Auto-detect format and parse from a file path."""
        uri = source_uri or f"file://{path.resolve()}"
        raw = path.read_bytes()
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            return self.parse_xlsx(raw, source_uri=uri)
        if suffix in (".tsv", "\t"):
            return self.parse_csv(raw, source_uri=uri, delimiter="\t")
        return self.parse_csv(raw, source_uri=uri)


def _digest(data: bytes) -> str:
    return f"sha256:{hashlib.sha256(data).hexdigest()}"
